#!/usr/bin/env python3
"""Analyze all xlsx files - lightweight version for large files."""

import openpyxl
import os
import json

EXTRACTED_DIR = "/home/z/my-project/extracted"
OUTPUT_FILE = "/home/z/my-project/scripts/analysis_result.json"

def analyze_workbook(filepath):
    """Analyze a single xlsx workbook - headers + sample rows only."""
    result = {
        "filename": os.path.basename(filepath),
        "file_size_mb": round(os.path.getsize(filepath) / (1024 * 1024), 2),
        "sheets": []
    }
    
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column
        }
        
        # Sample first 3 rows only
        sample_rows = []
        for i, row in enumerate(ws.iter_rows(max_row=3, values_only=False)):
            row_data = []
            for cell in row:
                if cell.value is not None:
                    val_str = str(cell.value)
                    if len(val_str) > 80:
                        val_str = val_str[:80] + "..."
                    row_data.append({
                        "col": cell.column_letter,
                        "value": val_str,
                        "type": type(cell.value).__name__
                    })
            if row_data:
                sample_rows.append(row_data)
        sheet_info["sample_rows"] = sample_rows
        sheet_info["headers"] = [c["col"] + ": " + c["value"] for c in sample_rows[0]] if sample_rows else []
        
        result["sheets"].append(sheet_info)
    
    wb.close()
    return result

def main():
    all_results = []
    
    xlsx_files = sorted([f for f in os.listdir(EXTRACTED_DIR) if f.endswith(".xlsx")])
    
    for fname in xlsx_files:
        fpath = os.path.join(EXTRACTED_DIR, fname)
        print(f"Analyzing: {fname}")
        try:
            info = analyze_workbook(fpath)
            all_results.append(info)
            print(f"  Sheets: {len(info['sheets'])}")
            for s in info["sheets"]:
                print(f"    - {s['name']}: {s['max_row']} rows x {s['max_column']} cols")
                if s['headers']:
                    print(f"      Headers: {s['headers'][:10]}")
        except Exception as e:
            print(f"  ERROR: {e}")
            all_results.append({"filename": fname, "error": str(e)})
    
    with open(OUTPUT_FILE, "w") as f:
        json.dump(all_results, f, indent=2, default=str)
    
    print(f"\nAnalysis saved to {OUTPUT_FILE}")
    
    # Consolidated summary
    total_rows = 0
    for r in all_results:
        if "error" not in r:
            for s in r["sheets"]:
                total_rows += s["max_row"]
    
    summary = {
        "total_files": len(xlsx_files),
        "total_data_rows": total_rows,
        "total_size_mb": round(sum(os.path.getsize(os.path.join(EXTRACTED_DIR, f)) for f in xlsx_files) / (1024*1024), 1)
    }
    print(f"\n{'='*60}")
    print("CONSOLIDATED SUMMARY")
    print(f"{'='*60}")
    print(f"Total files: {summary['total_files']}")
    print(f"Total data rows (incl. header): {total_rows}")
    print(f"Total size: {summary['total_size_mb']} MB")
    
    # Write summary separately
    with open("/home/z/my-project/scripts/summary.json", "w") as f:
        json.dump(summary, f, indent=2)

if __name__ == "__main__":
    main()
